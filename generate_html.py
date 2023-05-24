from bs4 import BeautifulSoup
import openpyxl
import random

# Load the Excel file
workbook = openpyxl.load_workbook('IMDB Movie Ratings.xlsx')
sheet = workbook['Top Rated Movies']

# Check if there are any rows (excluding the header row) in the sheet
if sheet.max_row > 1:
    # Get a random row from the sheet (excluding the header row)
    random_row = random.choice(list(sheet.iter_rows(min_row=2, values_only=True)))

    # Extract movie details from the random row
    rank = random_row[0]
    name = random_row[1]
    year = random_row[2]
    rating = random_row[3]
    image = random_row[4]
    link = random_row[5]

    print(link)
  
 #Create HTML code
    html = "<html>\n<head>\n<link rel='stylesheet' type='text/css' href='styles.css'>\n</head>\n<body>\n"
    html += '<h1>Random Movie Recommendation</h1>\n';
    html += "<div class='container'>\n"
    html += f"<div class='movie-image'><img src='{image}' alt='Movie Poster'></div>\n"
    html += "<div class='movie-info'>\n"
    html += f"<p class='rank'><strong>Rank:</strong> {rank}</p>\n"
    html += f"<p class='name'><strong>Name:</strong> {name}</p>\n"
    html += f"<p class='year'><strong>Year of Release:</strong> {year}</p>\n"
    html += f"<p class='rating'><strong>IMDB Rating:</strong> {rating}</p>\n"
    html += "<button class='info-button' onclick='window.open(\"" + link + "\", \"_blank\")'>More Info</button>\n"
    html += "</div>\n"  # Added closing div tag for 'movie-info'
    html += "</div>\n"  # Added closing div tag for 'container'
    html += "<button class='try-again-button' onclick='window.location.reload()' >Try Again</button>\n"  # Added button with class 'try-again-button'
    html += "</body>\n</html>"
    # Save the HTML code to a file
    with open('random_movie.html', 'w') as file:
        file.write(html)
else:
    print("No movie data available.")


