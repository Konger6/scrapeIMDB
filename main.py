
from bs4 import BeautifulSoup
import requests
import openpyxl

# Create a new Excel workbook
excel = openpyxl.Workbook()

# Print the sheet names (currently empty)
print(excel.sheetnames)

# Get the active sheet and set its title
sheet = excel.active
sheet.title = 'Top Rated Movies'

# Print the sheet names (should now show 'Top Rated Movies')
print(excel.sheetnames)

# Append header row to the sheet
sheet.append(['Movie Rank', 'Movie Name', 'Year of Release', 'IMDB Rating', 'Movie Image'])

try:
    # Send a GET request to the IMDb website
    source = requests.get('https://www.imdb.com/chart/top')

    # Raise an exception if the response is not successful
    source.raise_for_status()

    # Create a BeautifulSoup object from the HTML source
    soup = BeautifulSoup(source.text, 'html.parser')

    # Find all movie rows in the table
    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    # Iterate over each movie row
    for movie in movies:
        # Extract movie details from different HTML elements
        name = movie.find('td', class_="titleColumn").a.text
        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
        year = movie.find('td', class_="titleColumn").span.text.strip('()')
        rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
        image_url = movie.find('td', class_="posterColumn").a.img['src']  # Extract movie image URL
        link = "https://www.imdb.com" + movie.find('td', class_="titleColumn").find('a')['href']

        # Append the movie details to the sheet
        sheet.append([rank, name, year, rating, image_url, link])

except Exception as e:
    # Print any exception that occurs during the web scraping process
    print(e)

# Save the workbook to an Excel file
excel.save('IMDB Movie Ratings.xlsx')

# Run generate_html.py to generate a new random movie HTML
import subprocess
subprocess.run(['python', 'generate_html.py'])